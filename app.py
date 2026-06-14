import os
import secrets
import json
import csv
import io
import time
import random
from datetime import datetime

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response, session, g
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from cryptography.fernet import Fernet
import markdown2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

_secret = os.environ.get('SECRET_KEY')
if not _secret:
    _key_file = os.path.join(os.path.dirname(__file__), '.secret_key')
    if os.path.exists(_key_file):
        with open(_key_file) as f:
            _secret = f.read().strip()
    else:
        _secret = secrets.token_hex(32)
        with open(_key_file, 'w') as f:
            f.write(_secret)
app.config['SECRET_KEY'] = _secret

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///sysadmin.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['TEMPLATES_AUTO_RELOAD'] = False
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB 上传限制
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = '请先登录'

# Fernet encryption key — generated once, stored in .fernet_key
KEY_FILE = os.path.join(os.path.dirname(__file__), '.fernet_key')
if os.path.exists(KEY_FILE):
    with open(KEY_FILE, 'rb') as f:
        FERNET_KEY = f.read()
else:
    FERNET_KEY = Fernet.generate_key()
    with open(KEY_FILE, 'wb') as f:
        f.write(FERNET_KEY)
fernet = Fernet(FERNET_KEY)


# ── Models ────────────────────────────────────────────────────────────────────

class User(UserMixin, db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    username      = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    is_admin      = db.Column(db.Boolean, default=False, nullable=False)
    is_active     = db.Column(db.Boolean, default=True, nullable=False)

    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw, method='pbkdf2:sha256')

    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)


class PasswordEntry(db.Model):
    id                 = db.Column(db.Integer, primary_key=True)
    name               = db.Column(db.String(200), nullable=False)
    username           = db.Column(db.String(200), default='')
    encrypted_password = db.Column(db.Text, nullable=False)
    host               = db.Column(db.String(200), default='')
    port               = db.Column(db.String(20),  default='')
    notes              = db.Column(db.Text, default='')
    category           = db.Column(db.String(100), default='未分类')
    region             = db.Column(db.String(100), default='')
    conn_type          = db.Column(db.String(20),  default='')
    click_count        = db.Column(db.Integer, default=0)
    pin_order          = db.Column(db.Integer, nullable=True)
    user_id            = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    workspace_id       = db.Column(db.Integer, db.ForeignKey('workspace.id'), nullable=True)
    created_at         = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at         = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def set_password(self, pw):
        self.encrypted_password = fernet.encrypt(pw.encode()).decode()

    def get_password(self):
        return fernet.decrypt(self.encrypted_password.encode()).decode()


class Workspace(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(100), nullable=False)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    max_pins   = db.Column(db.Integer, default=5, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class PasswordOption(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    opt_type     = db.Column(db.String(20), nullable=False)  # category / region / conn_type
    value        = db.Column(db.String(100), nullable=False)
    user_id      = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    workspace_id = db.Column(db.Integer, db.ForeignKey('workspace.id'), nullable=True)



class Note(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    title        = db.Column(db.String(200), nullable=False, default='新建笔记')
    content      = db.Column(db.Text, default='')
    parent_id    = db.Column(db.Integer, db.ForeignKey('note.id'), nullable=True)
    sort_order   = db.Column(db.Integer, default=0)
    user_id      = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    workspace_id = db.Column(db.Integer, db.ForeignKey('workspace.id'), nullable=True)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at   = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


@login_manager.user_loader
def load_user(uid):
    return User.query.get(int(uid))


def ws_id():
    return session.get('ws_id')


def get_opts(opt_type):
    return [o.value for o in PasswordOption.query.filter_by(
        user_id=current_user.id, workspace_id=ws_id(), opt_type=opt_type
    ).order_by(PasswordOption.value).all()]


def ensure_opt(opt_type, value):
    if not value:
        return
    if not PasswordOption.query.filter_by(
        user_id=current_user.id, workspace_id=ws_id(), opt_type=opt_type, value=value
    ).first():
        db.session.add(PasswordOption(
            opt_type=opt_type, value=value,
            user_id=current_user.id, workspace_id=ws_id()
        ))


@app.after_request
def set_security_headers(response):
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' cdn.jsdelivr.net cdnjs.cloudflare.com 'unsafe-inline'; "
        "style-src 'self' cdn.jsdelivr.net 'unsafe-inline'; "
        "font-src 'self' cdn.jsdelivr.net data:; "
        "img-src 'self' data: blob:; "
        "connect-src 'self'; "
        "frame-ancestors 'self';"
    )
    return response


@app.before_request
def load_workspace():
    g.workspace = None
    g.workspaces = []
    if current_user.is_authenticated:
        wsid = session.get('ws_id')
        ws = Workspace.query.filter_by(id=wsid, user_id=current_user.id).first() if wsid else None
        if not ws:
            ws = Workspace.query.filter_by(user_id=current_user.id).order_by(Workspace.id).first()
            if ws:
                session['ws_id'] = ws.id
        g.workspace = ws
        g.workspaces = Workspace.query.filter_by(user_id=current_user.id).order_by(Workspace.id).all()


# ── Login attempt tracking ────────────────────────────────────────────────────

_login_attempts = {}  # username -> {'count': int, 'locked_until': float}
_ip_attempts    = {}  # ip       -> {'count': int, 'locked_until': float}

def _attempt_info(username):
    return _login_attempts.get(username, {'count': 0, 'locked_until': 0.0})

def _ip_info(ip):
    return _ip_attempts.get(ip, {'count': 0, 'locked_until': 0.0})

def _record_fail(username):
    info = dict(_attempt_info(username))
    info['count'] += 1
    c = info['count']
    if c >= 5:
        mins = 2 ** (c - 5)          # 5→1min 6→2min 7→4min 8→8min …
        info['locked_until'] = time.time() + min(mins, 1440) * 60
    _login_attempts[username] = info
    return info

def _record_ip_fail(ip):
    info = dict(_ip_info(ip))
    info['count'] += 1
    c = info['count']
    if c >= 20:
        mins = 2 ** ((c - 20) // 5)  # 20-24: 1min, 25-29: 2min, 30-34: 4min…
        info['locked_until'] = time.time() + min(mins, 1440) * 60
    _ip_attempts[ip] = info
    return info

def _reset_attempts(username):
    _login_attempts.pop(username, None)

def _reset_ip_attempts(ip):
    _ip_attempts.pop(ip, None)

def _new_captcha():
    if random.randint(0, 1):
        a, b, c = random.randint(2, 9), random.randint(2, 9), random.randint(2, 6)
        q, ans = f"({a} + {b}) × {c}", (a + b) * c
    else:
        a, b = random.randint(3, 9), random.randint(3, 9)
        c = random.randint(1, a * b - 1)
        q, ans = f"{a} × {b} - {c}", a * b - c
    session['captcha_ans'] = ans
    session['captcha_q']   = q + " = ?"
    return session['captcha_q']


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    users = User.query.order_by(User.username).all()

    # 确保验证码已生成（GET 时也预生成，避免第一次 POST 时 session 里没有）
    if 'captcha_ans' not in session:
        _new_captcha()
    captcha_q   = session.get('captcha_q')
    show_captcha = session.get('login_needs_captcha', False)

    if request.method == 'POST':
        username  = request.form.get('username', '').strip()
        password  = request.form.get('password', '')
        client_ip = request.remote_addr
        info      = _attempt_info(username)

        # 0. IP 是否被封锁
        ip_remaining = _ip_info(client_ip)['locked_until'] - time.time()
        if ip_remaining > 0:
            mins = max(1, int(ip_remaining // 60) + 1)
            flash(f'该 IP 登录尝试过多，请 {mins} 分钟后再试', 'danger')
            return render_template('login.html', users=users,
                                   captcha_q=captcha_q, show_captcha=show_captcha)

        # 1. 账户是否被锁定
        remaining = info['locked_until'] - time.time()
        if remaining > 0:
            mins = max(1, int(remaining // 60) + 1)
            flash(f'账户已锁定，请 {mins} 分钟后再试', 'danger')
            return render_template('login.html', users=users,
                                   captcha_q=captcha_q, show_captcha=show_captcha)

        # 2. 第 3 次起需要验证码
        if info['count'] >= 3:
            user_ans    = request.form.get('captcha', '').strip()
            correct_ans = session.pop('captcha_ans', None)
            session.pop('captcha_q', None)
            captcha_q = _new_captcha()
            show_captcha = True
            session['login_needs_captcha'] = True
            try:
                ok = int(user_ans) == correct_ans
            except (ValueError, TypeError):
                ok = False
            if not ok:
                flash('验证码错误', 'danger')
                return render_template('login.html', users=users,
                                       captcha_q=captcha_q, show_captcha=True)

        # 3. 验证密码
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            if not user.is_active:
                flash('该帐户已被禁用', 'danger')
            else:
                _reset_attempts(username)
                _reset_ip_attempts(client_ip)
                session.pop('login_needs_captcha', None)
                session.pop('captcha_ans', None)
                session.pop('captcha_q', None)
                login_user(user, remember=False)
                next_url = request.args.get('next', '')
                if next_url and next_url.startswith('/') and not next_url.startswith('//'):
                    return redirect(next_url)
                return redirect(url_for('index'))
        else:
            new_info = _record_fail(username)
            _record_ip_fail(client_ip)
            c = new_info['count']
            if c >= 5:
                mins = max(1, int((new_info['locked_until'] - time.time()) // 60) + 1)
                flash(f'密码连续错误 {c} 次，账户已锁定 {mins} 分钟', 'danger')
            else:
                flash(f'用户名或密码错误（第 {c} 次）', 'danger')
            if c >= 3:
                session['login_needs_captcha'] = True
                show_captcha = True
                if 'captcha_ans' not in session:
                    captcha_q = _new_captcha()

    return render_template('login.html', users=users,
                           captcha_q=captcha_q, show_captcha=show_captcha)


@app.route('/logout', methods=['POST'])
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    return redirect(url_for('passwords_list'))


@app.route('/search')
@login_required
def search():
    q = request.args.get('q', '').strip()
    if not q:
        return redirect(url_for('passwords_list'))
    passwords = PasswordEntry.query.filter_by(user_id=current_user.id, workspace_id=ws_id()).filter(
        PasswordEntry.name.contains(q) |
        PasswordEntry.host.contains(q) |
        PasswordEntry.username.contains(q) |
        PasswordEntry.category.contains(q) |
        PasswordEntry.region.contains(q) |
        PasswordEntry.conn_type.contains(q) |
        PasswordEntry.notes.contains(q)
    ).order_by(PasswordEntry.updated_at.desc()).all()
    notes = Note.query.filter_by(user_id=current_user.id, workspace_id=ws_id()).filter(
        Note.title.contains(q) | Note.content.contains(q)
    ).order_by(Note.updated_at.desc()).all()
    return render_template('search.html', q=q, passwords=passwords, notes=notes)


# ── Notes ─────────────────────────────────────────────────────────────────────

@app.route('/notes')
@login_required
def notes_main():
    return render_template('notes/main.html')


@app.route('/notes/api/tree')
@login_required
def notes_api_tree():
    items = Note.query.filter_by(user_id=current_user.id, workspace_id=ws_id())\
                      .order_by(Note.sort_order, Note.created_at).all()
    return jsonify([{'id': n.id, 'title': n.title, 'parent_id': n.parent_id} for n in items])


@app.route('/notes/api/<int:id>')
@login_required
def notes_api_get(id):
    note = Note.query.filter_by(id=id, user_id=current_user.id, workspace_id=ws_id()).first_or_404()
    return jsonify({
        'id': note.id, 'title': note.title, 'content': note.content,
        'parent_id': note.parent_id,
        'updated_at': note.updated_at.strftime('%m-%d %H:%M')
    })


@app.route('/notes/api/new', methods=['POST'])
@login_required
def notes_api_new():
    data = request.get_json(silent=True) or {}
    parent_id = data.get('parent_id')
    if parent_id:
        if not Note.query.filter_by(id=parent_id, user_id=current_user.id, workspace_id=ws_id()).first():
            return jsonify({'error': 'invalid parent'}), 400
    note = Note(title='新建笔记', content='', parent_id=parent_id, user_id=current_user.id, workspace_id=ws_id())
    db.session.add(note)
    db.session.commit()
    return jsonify({'id': note.id, 'title': note.title, 'parent_id': note.parent_id})


@app.route('/notes/api/<int:id>/save', methods=['POST'])
@login_required
def notes_api_save(id):
    note = Note.query.filter_by(id=id, user_id=current_user.id, workspace_id=ws_id()).first_or_404()
    data = request.get_json(silent=True) or {}
    note.title      = (data.get('title') or '新建笔记').strip()
    note.content    = data.get('content', '')
    note.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'ok': True, 'updated_at': note.updated_at.strftime('%m-%d %H:%M')})


@app.route('/notes/api/<int:id>/move', methods=['POST'])
@login_required
def notes_api_move(id):
    note = Note.query.filter_by(id=id, user_id=current_user.id, workspace_id=ws_id()).first_or_404()
    data = request.get_json(silent=True) or {}
    new_parent_id = data.get('parent_id')  # None = root
    if new_parent_id:
        parent = Note.query.filter_by(id=new_parent_id, user_id=current_user.id, workspace_id=ws_id()).first_or_404()
        # prevent circular: walk ancestors of new parent
        cur = parent
        while cur:
            if cur.id == id:
                return jsonify({'error': 'circular'}), 400
            cur = Note.query.get(cur.parent_id) if cur.parent_id else None
    note.parent_id = new_parent_id
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/notes/api/<int:id>/delete', methods=['POST'])
@login_required
def notes_api_delete(id):
    Note.query.filter_by(id=id, user_id=current_user.id, workspace_id=ws_id()).first_or_404()
    def _del(nid):
        for child in Note.query.filter_by(parent_id=nid).all():
            _del(child.id)
        Note.query.filter_by(id=nid).delete()
    _del(id)
    db.session.commit()
    return jsonify({'ok': True})



# ── Password Vault ────────────────────────────────────────────────────────────

@app.route('/passwords')
@login_required
def passwords_list():
    category  = request.args.get('category', '')
    region    = request.args.get('region', '')
    q         = request.args.get('q', '')
    query     = PasswordEntry.query.filter_by(user_id=current_user.id, workspace_id=ws_id())
    if category:
        query = query.filter_by(category=category)
    if region:
        query = query.filter_by(region=region)
    if q:
        query = query.filter(
            PasswordEntry.name.contains(q) |
            PasswordEntry.host.contains(q) |
            PasswordEntry.username.contains(q) |
            PasswordEntry.category.contains(q) |
            PasswordEntry.region.contains(q) |
            PasswordEntry.conn_type.contains(q) |
            PasswordEntry.notes.contains(q)
        )
    passwords  = query.order_by(
        db.case((PasswordEntry.pin_order == None, 1), else_=0),
        PasswordEntry.pin_order.asc(),
        PasswordEntry.updated_at.desc()
    ).all()
    categories = db.session.query(PasswordEntry.category, db.func.count(PasswordEntry.id))\
                           .filter_by(user_id=current_user.id, workspace_id=ws_id())\
                           .group_by(PasswordEntry.category).all()
    regions = db.session.query(PasswordEntry.region, db.func.count(PasswordEntry.id))\
                        .filter_by(user_id=current_user.id, workspace_id=ws_id())\
                        .filter(PasswordEntry.region != '')\
                        .group_by(PasswordEntry.region).all()
    return render_template('passwords/list.html',
                           passwords=passwords, categories=categories, regions=regions,
                           current_category=category, current_region=region, search=q)


@app.route('/passwords/new', methods=['GET', 'POST'])
@login_required
def password_new():
    if request.method == 'POST':
        is_json = request.is_json
        data    = request.get_json(silent=True) or {} if is_json else request.form
        name     = (data.get('name') or '').strip()
        pw       = data.get('password') or ''
        category = (data.get('category') or data.get('new_category') or '未分类').strip()
        region   = (data.get('region') if data.get('region') not in ('', '__new__', None) else data.get('new_region') or '').strip()
        if not name:
            if is_json:
                return jsonify({'error': '名称不能为空'}), 400
            flash('名称不能为空', 'danger')
        else:
            entry = PasswordEntry(
                name=name,
                username=(data.get('username') or '').strip(),
                host=(data.get('host') or '').strip(),
                port=(data.get('port') or '').strip(),
                notes=data.get('notes') or '',
                category=category,
                region=region,
                conn_type=(data.get('conn_type') or '').strip(),
                user_id=current_user.id,
                workspace_id=ws_id()
            )
            entry.set_password(pw)
            db.session.add(entry)
            ensure_opt('category', category)
            ensure_opt('region', region)
            ensure_opt('conn_type', (data.get('conn_type') or '').strip())
            db.session.commit()
            if is_json:
                return jsonify({'ok': True, 'id': entry.id})
            return redirect(url_for('passwords_list'))
    cats = get_opts('category')
    regs = get_opts('region')
    return render_template('passwords/edit.html', entry=None, categories=cats, regions=regs)


@app.route('/passwords/<int:id>')
@login_required
def password_view(id):
    entry = PasswordEntry.query.filter_by(id=id, user_id=current_user.id, workspace_id=ws_id()).first_or_404()
    return render_template('passwords/view.html', entry=entry)



@app.route('/passwords/<int:id>/reveal')
@login_required
def password_reveal(id):
    entry = PasswordEntry.query.filter_by(id=id, user_id=current_user.id, workspace_id=ws_id()).first_or_404()
    return jsonify({'password': entry.get_password()})


@app.route('/passwords/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def password_edit(id):
    entry = PasswordEntry.query.filter_by(id=id, user_id=current_user.id, workspace_id=ws_id()).first_or_404()
    if request.method == 'POST':
        is_json = request.is_json
        data    = request.get_json(silent=True) or {} if is_json else request.form
        name     = (data.get('name') or '').strip()
        category = (data.get('category') or data.get('new_category') or '未分类').strip()
        if not name:
            if is_json:
                return jsonify({'error': '名称不能为空'}), 400
            flash('名称不能为空', 'danger')
        else:
            region           = (data.get('region') if data.get('region') not in ('', '__new__', None) else data.get('new_region') or '').strip()
            entry.name       = name
            entry.username   = (data.get('username') or '').strip()
            entry.host       = (data.get('host') or '').strip()
            entry.port       = (data.get('port') or '').strip()
            entry.notes      = data.get('notes') or entry.notes
            entry.category   = category
            entry.region     = region
            entry.conn_type  = (data.get('conn_type') or '').strip()
            entry.updated_at = datetime.utcnow()
            new_pw = data.get('password') or ''
            if new_pw:
                entry.set_password(new_pw)
            ensure_opt('category', entry.category)
            ensure_opt('region', entry.region)
            ensure_opt('conn_type', entry.conn_type)
            db.session.commit()
            if is_json:
                return jsonify({'ok': True})
            flash('已更新', 'success')
            return redirect(url_for('password_view', id=entry.id))
    cats = get_opts('category')
    regs = get_opts('region')
    return render_template('passwords/edit.html', entry=entry, categories=cats, regions=regs)


@app.route('/passwords/<int:id>/delete', methods=['POST'])
@login_required
def password_delete(id):
    entry = PasswordEntry.query.filter_by(id=id, user_id=current_user.id, workspace_id=ws_id()).first_or_404()
    db.session.delete(entry)
    db.session.commit()
    if request.is_json:
        return jsonify({'ok': True})
    flash('已删除', 'success')
    return redirect(url_for('passwords_list'))


@app.route('/passwords/<int:id>/pin', methods=['POST'])
@login_required
def password_pin(id):
    entry = PasswordEntry.query.filter_by(id=id, user_id=current_user.id, workspace_id=ws_id()).first_or_404()
    if entry.pin_order is not None:
        # 已置顶 → 取消，低于它的往上补位
        removed = entry.pin_order
        entry.pin_order = None
        PasswordEntry.query.filter(
            PasswordEntry.user_id == current_user.id,
            PasswordEntry.workspace_id == ws_id(),
            PasswordEntry.pin_order > removed
        ).update({'pin_order': PasswordEntry.pin_order - 1})
    else:
        ws = Workspace.query.get(ws_id())
        max_pins = ws.max_pins if ws else 5
        PasswordEntry.query.filter(
            PasswordEntry.user_id == current_user.id,
            PasswordEntry.workspace_id == ws_id(),
            PasswordEntry.pin_order.isnot(None)
        ).update({'pin_order': PasswordEntry.pin_order + 1})
        PasswordEntry.query.filter(
            PasswordEntry.user_id == current_user.id,
            PasswordEntry.workspace_id == ws_id(),
            PasswordEntry.pin_order > max_pins
        ).update({'pin_order': None})
        entry.pin_order = 1
    db.session.commit()
    return ('', 204)


# ── User Management ───────────────────────────────────────────────────────────

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('仅管理员可访问', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


@app.route('/users')
@login_required
@admin_required
def users_list():
    users = User.query.order_by(User.id).all()
    return render_template('users/list.html', users=users)


@app.route('/users/new', methods=['GET', 'POST'])
@login_required
@admin_required
def user_new():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        if not username or not password:
            flash('用户名和密码不能为空', 'danger')
        elif User.query.filter_by(username=username).first():
            flash('用户名已存在', 'danger')
        else:
            u = User(username=username)
            u.set_password(password)
            db.session.add(u)
            db.session.commit()
            flash(f'用户 {username} 已创建', 'success')
            return redirect(url_for('users_list'))
    return render_template('users/edit.html', user=None)


@app.route('/users/<int:id>/reset', methods=['GET', 'POST'])
@login_required
@admin_required
def user_reset(id):
    user = User.query.get_or_404(id)
    if request.method == 'POST':
        pw = request.form.get('password', '')
        if not pw:
            flash('密码不能为空', 'danger')
        else:
            user.set_password(pw)
            db.session.commit()
            flash(f'{user.username} 密码已重置', 'success')
            return redirect(url_for('users_list'))
    return render_template('users/edit.html', user=user, reset=True)


@app.route('/users/<int:id>/toggle_active', methods=['POST'])
@login_required
@admin_required
def user_toggle_active(id):
    user = User.query.get_or_404(id)
    user.is_active = not user.is_active
    db.session.commit()
    state = '启用' if user.is_active else '禁用'
    flash(f'帐户 {user.username} 已{state}', 'success')
    return redirect(url_for('users_list'))


@app.route('/users/<int:id>/toggle_admin', methods=['POST'])
@login_required
@admin_required
def user_toggle_admin(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id and user.is_admin:
        flash('不能取消自己的管理员权限', 'danger')
        return redirect(url_for('users_list'))
    user.is_admin = not user.is_admin
    db.session.commit()
    state = '授予' if user.is_admin else '取消'
    flash(f'已{state} {user.username} 的管理员权限', 'success')
    return redirect(url_for('users_list'))


@app.route('/users/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def user_delete(id):
    user = User.query.get_or_404(id)
    if user.username == 'admin':
        flash('不能删除 admin 账号', 'danger')
    else:
        db.session.delete(user)
        db.session.commit()
        flash(f'用户 {user.username} 已删除', 'success')
    return redirect(url_for('users_list'))


# ── Options Management ───────────────────────────────────────────────────────

@app.route('/options')
@login_required
def options_list():
    cats  = PasswordOption.query.filter_by(user_id=current_user.id, workspace_id=ws_id(), opt_type='category').order_by(PasswordOption.value).all()
    regs  = PasswordOption.query.filter_by(user_id=current_user.id, workspace_id=ws_id(), opt_type='region').order_by(PasswordOption.value).all()
    types = PasswordOption.query.filter_by(user_id=current_user.id, workspace_id=ws_id(), opt_type='conn_type').order_by(PasswordOption.value).all()
    return render_template('options/list.html', cats=cats, regs=regs, types=types)


@app.route('/options/add', methods=['POST'])
@login_required
def option_add():
    opt_type = request.form.get('opt_type', '').strip()
    value    = request.form.get('value', '').strip()
    if opt_type not in ('category', 'region', 'conn_type') or not value:
        flash('参数错误', 'danger')
        return redirect(url_for('options_list'))
    if not PasswordOption.query.filter_by(
        user_id=current_user.id, workspace_id=ws_id(), opt_type=opt_type, value=value
    ).first():
        db.session.add(PasswordOption(opt_type=opt_type, value=value,
                                      user_id=current_user.id, workspace_id=ws_id()))
        db.session.commit()
    return redirect(url_for('options_list'))


@app.route('/options/<int:id>/delete', methods=['POST'])
@login_required
def option_delete(id):
    opt = PasswordOption.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    db.session.delete(opt)
    db.session.commit()
    return redirect(url_for('options_list'))


# ── Workspace ────────────────────────────────────────────────────────────────

@app.route('/workspaces')
@login_required
def workspaces_list():
    workspaces = Workspace.query.filter_by(user_id=current_user.id).order_by(Workspace.id).all()
    return render_template('workspaces/list.html', workspaces=workspaces)


@app.route('/workspaces/create', methods=['POST'])
@login_required
def workspace_create():
    name = (request.form.get('name') or '').strip()
    if not name:
        flash('名称不能为空', 'danger')
        return redirect(url_for('workspaces_list'))
    ws = Workspace(name=name, user_id=current_user.id)
    db.session.add(ws)
    db.session.commit()
    session['ws_id'] = ws.id
    flash(f'已创建并切换到「{name}」', 'success')
    return redirect(url_for('workspaces_list'))


@app.route('/workspaces/<int:id>/switch', methods=['POST'])
@login_required
def workspace_switch(id):
    ws = Workspace.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    session['ws_id'] = ws.id
    return redirect(url_for('passwords_list'))


@app.route('/workspaces/<int:id>/rename', methods=['POST'])
@login_required
def workspace_rename(id):
    ws = Workspace.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    name = (request.form.get('name') or '').strip()
    if not name:
        flash('名称不能为空', 'danger')
        return redirect(url_for('workspaces_list'))
    ws.name = name
    db.session.commit()
    flash('已重命名', 'success')
    return redirect(url_for('workspaces_list'))


@app.route('/workspaces/<int:id>/set_max_pins', methods=['POST'])
@login_required
def workspace_set_max_pins(id):
    ws = Workspace.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    try:
        val = int(request.form.get('max_pins', 5))
        ws.max_pins = max(1, min(20, val))
        db.session.commit()
    except (ValueError, TypeError):
        pass
    return redirect(url_for('workspaces_list'))


@app.route('/workspaces/<int:id>/delete', methods=['POST'])
@login_required
def workspace_delete(id):
    ws = Workspace.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    if Workspace.query.filter_by(user_id=current_user.id).count() <= 1:
        flash('至少保留一个数据库', 'danger')
        return redirect(url_for('workspaces_list'))
    pw_count = PasswordEntry.query.filter_by(workspace_id=id).count()
    note_count = Note.query.filter_by(workspace_id=id).count()
    if pw_count + note_count > 0:
        flash(f'该数据库还有 {pw_count} 条密码、{note_count} 条笔记，请先清空再删除', 'danger')
        return redirect(url_for('workspaces_list'))
    if session.get('ws_id') == id:
        other = Workspace.query.filter_by(user_id=current_user.id).filter(Workspace.id != id).first()
        session['ws_id'] = other.id if other else None
    db.session.delete(ws)
    db.session.commit()
    flash('已删除', 'success')
    return redirect(url_for('workspaces_list'))


# ── Data Import / Export ──────────────────────────────────────────────────────

@app.route('/data')
@login_required
@admin_required
def data_manager():
    pw_count   = PasswordEntry.query.filter_by(user_id=current_user.id, workspace_id=ws_id()).count()
    note_count = Note.query.filter_by(user_id=current_user.id, workspace_id=ws_id()).count()
    return render_template('data.html', pw_count=pw_count, note_count=note_count)


@app.route('/data/export/passwords')
@login_required
@admin_required
def export_passwords():
    entries = PasswordEntry.query.filter_by(user_id=current_user.id, workspace_id=ws_id()).order_by(PasswordEntry.id).all()
    rows = []
    for e in entries:
        rows.append({
            'name': e.name, 'username': e.username,
            'password': e.get_password(),
            'host': e.host, 'port': e.port,
            'conn_type': e.conn_type, 'category': e.category,
            'region': e.region, 'notes': e.notes,
        })
    data = json.dumps(rows, ensure_ascii=False, indent=2)
    return Response(data, mimetype='application/json',
                    headers={'Content-Disposition': f'attachment;filename=passwords_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'})


@app.route('/passwords/export/xlsx')
@login_required
def passwords_export_xlsx():
    entries = PasswordEntry.query.filter_by(user_id=current_user.id, workspace_id=ws_id()).order_by(PasswordEntry.id).all()
    wb = openpyxl.Workbook()
    ws_ = wb.active
    ws_.title = '密码'
    headers = ['名称', '用户名', '密码', '主机/IP', '端口', '连接类型', '分类', '地区', '备注']
    fields  = ['name','username','password','host','port','conn_type','category','region','notes']
    hdr_fill = PatternFill('solid', fgColor='2D333B')
    hdr_font = Font(bold=True, color='E6EDF3')
    for ci, h in enumerate(headers, 1):
        cell = ws_.cell(1, ci, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
    for ri, e in enumerate(entries, 2):
        row_data = [e.name, e.username, e.get_password(), e.host, e.port, e.conn_type, e.category, e.region, e.notes]
        for ci, v in enumerate(row_data, 1):
            ws_.cell(ri, ci, v or '')
    col_widths = [20, 18, 18, 20, 8, 12, 14, 12, 30]
    for ci, w in enumerate(col_widths, 1):
        ws_.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment;filename=passwords_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'})


@app.route('/passwords/import/xlsx', methods=['POST'])
@login_required
def passwords_import_xlsx():
    f = request.files.get('file')
    if not f or not f.filename.endswith('.xlsx'):
        return jsonify({'error': '请上传 .xlsx 文件'}), 400
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws_ = wb.active
        headers = [str(c.value).strip() if c.value else '' for c in next(ws_.iter_rows(min_row=1, max_row=1))]
        col_map = {'名称':'name','用户名':'username','密码':'password','主机/IP':'host',
                   '端口':'port','连接类型':'conn_type','分类':'category','地区':'region','备注':'notes'}
        idx = {col_map[h]: i for i, h in enumerate(headers) if h in col_map}
        added = 0
        for row in ws_.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            def g(k): return str(row[idx[k]]).strip() if k in idx and row[idx[k]] is not None else ''
            pw = g('password')
            if not pw:
                continue
            e = PasswordEntry(name=g('name') or '导入', username=g('username'),
                              host=g('host'), port=g('port'), conn_type=g('conn_type'),
                              category=g('category') or '未分类', region=g('region'), notes=g('notes'),
                              user_id=current_user.id, workspace_id=ws_id())
            e.set_password(pw)
            db.session.add(e)
            added += 1
        db.session.commit()
        return jsonify({'ok': True, 'added': added})
    except Exception as ex:
        return jsonify({'error': str(ex)}), 400


@app.route('/data/export/notes')
@login_required
@admin_required
def export_notes():
    notes = Note.query.filter_by(user_id=current_user.id, workspace_id=ws_id()).order_by(Note.id).all()
    rows = [{'id': n.id, 'title': n.title, 'content': n.content, 'parent_id': n.parent_id} for n in notes]
    data = json.dumps(rows, ensure_ascii=False, indent=2)
    return Response(data, mimetype='application/json',
                    headers={'Content-Disposition': f'attachment;filename=notes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'})


@app.route('/data/export/notes/txt')
@login_required
@admin_required
def export_notes_txt():
    notes = Note.query.filter_by(user_id=current_user.id, workspace_id=ws_id()).order_by(Note.id).all()
    id_map = {n.id: n for n in notes}
    roots = [n for n in notes if not n.parent_id or n.parent_id not in id_map]

    lines = []
    def write_note(note, depth=0):
        indent = '  ' * depth
        lines.append(f"{indent}{'=' * max(1, 40 - len(indent))}")
        lines.append(f"{indent}{note.title}")
        lines.append(f"{indent}{'=' * max(1, 40 - len(indent))}")
        if note.content and note.content.strip():
            for line in note.content.splitlines():
                lines.append(f"{indent}{line}")
        lines.append('')
        for child in [n for n in notes if n.parent_id == note.id]:
            write_note(child, depth + 1)

    for root in roots:
        write_note(root)

    text = '\n'.join(lines)
    filename = f'notes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
    return Response(text, mimetype='text/plain; charset=utf-8',
                    headers={'Content-Disposition': f'attachment;filename={filename}'})


@app.route('/data/export/notes/md')
@login_required
@admin_required
def export_notes_md():
    notes = Note.query.filter_by(user_id=current_user.id, workspace_id=ws_id()).order_by(Note.id).all()
    id_map = {n.id: n for n in notes}
    roots = [n for n in notes if not n.parent_id or n.parent_id not in id_map]

    lines = []
    def write_note(note, depth=1):
        prefix = '#' * min(depth, 6)
        lines.append(f"{prefix} {note.title}")
        lines.append('')
        if note.content and note.content.strip():
            lines.append(note.content.strip())
            lines.append('')
        for child in [n for n in notes if n.parent_id == note.id]:
            write_note(child, depth + 1)

    for root in roots:
        write_note(root)

    text = '\n'.join(lines)
    filename = f'notes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.md'
    return Response(text, mimetype='text/markdown; charset=utf-8',
                    headers={'Content-Disposition': f'attachment;filename={filename}'})


@app.route('/data/import/passwords', methods=['POST'])
@login_required
@admin_required
def import_passwords():
    f = request.files.get('file')
    if not f:
        flash('请选择文件', 'danger')
        return redirect(url_for('data_manager'))
    try:
        rows = json.loads(f.read().decode('utf-8'))
        added = 0
        for r in rows:
            e = PasswordEntry(
                name      = (r.get('name') or '').strip(),
                username  = (r.get('username') or '').strip(),
                host      = (r.get('host') or '').strip(),
                port      = (r.get('port') or '').strip(),
                conn_type = (r.get('conn_type') or '').strip(),
                category  = (r.get('category') or '未分类').strip(),
                region    = (r.get('region') or '').strip(),
                notes     = r.get('notes') or '',
                user_id      = current_user.id,
                workspace_id = ws_id(),
            )
            e.set_password(r.get('password') or '')
            db.session.add(e)
            added += 1
        db.session.commit()
        flash(f'成功导入 {added} 条密码记录', 'success')
    except Exception as ex:
        db.session.rollback()
        flash(f'导入失败：{ex}', 'danger')
    return redirect(url_for('data_manager'))


@app.route('/data/import/notes', methods=['POST'])
@login_required
@admin_required
def import_notes():
    f = request.files.get('file')
    if not f:
        flash('请选择文件', 'danger')
        return redirect(url_for('data_manager'))
    try:
        rows = json.loads(f.read().decode('utf-8'))
        # 先建立 old_id → new_id 映射以保留父子关系
        id_map = {}
        added = 0
        for r in rows:
            parent_new = id_map.get(r.get('parent_id'))
            n = Note(
                title     = (r.get('title') or '新建笔记').strip(),
                content   = r.get('content') or '',
                parent_id    = parent_new,
                user_id      = current_user.id,
                workspace_id = ws_id(),
            )
            db.session.add(n)
            db.session.flush()
            id_map[r.get('id')] = n.id
            added += 1
        db.session.commit()
        flash(f'成功导入 {added} 条笔记', 'success')
    except Exception as ex:
        db.session.rollback()
        flash(f'导入失败：{ex}', 'danger')
    return redirect(url_for('data_manager'))


@app.route('/data/import/notes/md', methods=['POST'])
@login_required
@admin_required
def import_notes_md():
    files = request.files.getlist('file')
    if not files:
        flash('请选择文件', 'danger')
        return redirect(url_for('data_manager'))
    try:
        added = 0
        for f in files:
            if not f.filename.endswith('.md'):
                continue
            title = f.filename[:-3].strip() or '导入笔记'
            content = f.read().decode('utf-8')
            n = Note(title=title, content=content,
                     user_id=current_user.id, workspace_id=ws_id())
            db.session.add(n)
            added += 1
        db.session.commit()
        flash(f'成功导入 {added} 条笔记', 'success')
    except Exception as ex:
        db.session.rollback()
        flash(f'导入失败：{ex}', 'danger')
    return redirect(url_for('data_manager'))


# ── Init & Run ────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    with app.app_context():
        # If note table has old schema (missing parent_id), drop and recreate
        try:
            with db.engine.connect() as conn:
                cols = {row[1] for row in conn.execute(db.text('PRAGMA table_info(note)')).fetchall()}
                if cols and 'parent_id' not in cols:
                    conn.execute(db.text('DROP TABLE note'))
                    conn.commit()
        except Exception:
            pass
        try:
            with db.engine.connect() as conn:
                cols = {row[1] for row in conn.execute(db.text('PRAGMA table_info(password_entry)')).fetchall()}
                if cols and 'conn_type' not in cols:
                    conn.execute(db.text("ALTER TABLE password_entry ADD COLUMN conn_type VARCHAR(20) DEFAULT ''"))
                    conn.commit()
        except Exception:
            pass
        try:
            with db.engine.connect() as conn:
                cols = {row[1] for row in conn.execute(db.text('PRAGMA table_info(password_entry)')).fetchall()}
                if cols and 'region' not in cols:
                    conn.execute(db.text("ALTER TABLE password_entry ADD COLUMN region VARCHAR(100) DEFAULT ''"))
                    conn.commit()
                if cols and 'click_count' not in cols:
                    conn.execute(db.text("ALTER TABLE password_entry ADD COLUMN click_count INTEGER DEFAULT 0"))
                    conn.commit()
                if cols and 'workspace_id' not in cols:
                    conn.execute(db.text("ALTER TABLE password_entry ADD COLUMN workspace_id INTEGER"))
                    conn.commit()
                if cols and 'pin_order' not in cols:
                    conn.execute(db.text("ALTER TABLE password_entry ADD COLUMN pin_order INTEGER"))
                    conn.commit()
        except Exception:
            pass
        try:
            with db.engine.connect() as conn:
                cols = {row[1] for row in conn.execute(db.text('PRAGMA table_info(note)')).fetchall()}
                if cols and 'workspace_id' not in cols:
                    conn.execute(db.text("ALTER TABLE note ADD COLUMN workspace_id INTEGER"))
                    conn.commit()
        except Exception:
            pass
        try:
            with db.engine.connect() as conn:
                cols = {row[1] for row in conn.execute(db.text('PRAGMA table_info(workspace)')).fetchall()}
                if cols and 'max_pins' not in cols:
                    conn.execute(db.text("ALTER TABLE workspace ADD COLUMN max_pins INTEGER DEFAULT 5"))
                    conn.commit()
        except Exception:
            pass
        try:
            with db.engine.connect() as conn:
                cols = {row[1] for row in conn.execute(db.text('PRAGMA table_info(user)')).fetchall()}
                if cols and 'is_admin' not in cols:
                    conn.execute(db.text("ALTER TABLE user ADD COLUMN is_admin BOOLEAN DEFAULT 0"))
                    conn.execute(db.text("UPDATE user SET is_admin=1 WHERE username='admin'"))
                    conn.commit()
                if cols and 'is_active' not in cols:
                    conn.execute(db.text("ALTER TABLE user ADD COLUMN is_active BOOLEAN DEFAULT 1"))
                    conn.commit()
        except Exception:
            pass
        db.create_all()
        if not User.query.first():
            admin = User(username='admin', is_admin=True)
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print('✓ 已创建默认用户: admin / admin123  (请登录后立即修改密码)')
        # Ensure existing admin user has is_admin=True
        admin_user = User.query.filter_by(username='admin').first()
        if admin_user and not admin_user.is_admin:
            admin_user.is_admin = True
            db.session.commit()
        # Ensure every user has at least one workspace; migrate orphan records
        for user in User.query.all():
            ws = Workspace.query.filter_by(user_id=user.id).order_by(Workspace.id).first()
            if not ws:
                ws = Workspace(name='默认', user_id=user.id)
                db.session.add(ws)
                db.session.flush()
            PasswordEntry.query.filter_by(user_id=user.id, workspace_id=None)\
                         .update({'workspace_id': ws.id})
            Note.query.filter_by(user_id=user.id, workspace_id=None)\
                .update({'workspace_id': ws.id})
            # Seed PasswordOption from existing entry values
            for entry in PasswordEntry.query.filter_by(user_id=user.id).all():
                for opt_type, val in [('category', entry.category), ('region', entry.region), ('conn_type', entry.conn_type)]:
                    if val and not PasswordOption.query.filter_by(
                        user_id=user.id, workspace_id=entry.workspace_id, opt_type=opt_type, value=val
                    ).first():
                        db.session.add(PasswordOption(
                            opt_type=opt_type, value=val,
                            user_id=user.id, workspace_id=entry.workspace_id
                        ))
        db.session.commit()
    port = int(os.environ.get('PORT', 5001))
    app.run(host='0.0.0.0', port=port, debug=False)
